from datetime import datetime
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from rtm_con.msg_flatten import flat_msg
from rtm_con.common_items import DataItem

class MsgExcel:
    msg_key: str = 'Msg'
    logtime_key: str = 'LogTime'
    prefixed_headers: list = [logtime_key, 'timestamp', msg_key]
    preset_formats: dict[str, str] = {
        logtime_key: 'yyyy-mm-dd hh:mm:ss.000',
        'timestamp': 'yyyy-mm-dd hh:mm:ss',
    }
    frozen_position: str = 'C2'  # Freeze first row and first two columns
    int_format: str = '0'
    float_format: str = '0.000'
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Logs"
        self.headers: list[str] = []
        self.headerpaths: dict[str, tuple[str, ...]] = {}
        self._update_headers(self.prefixed_headers)  # Predefine some common headers
        self.current_row: int = 2 # as the first row for headers
    
    def write_line(self, line_dict: flat_msg, *, pathdict: dict = {}) -> None:
        self._update_headers(line_dict.keys(), pathdict=pathdict)
        return self._write_line(line_dict)
    
    def save(self, path: str) -> None:
        self._presave_formatting()
        return self.wb.save(path)
    
    def get_column(self, col_name: str) -> ColumnDimension | None:
        if col_name in self.headers:
            col_letter = get_column_letter(self.headers.index(col_name) + 1)
            return self.ws.column_dimensions[col_letter]
        return None
    
    def _update_headers(self, new_headers: list, *, pathdict: dict = {}) -> None:
        for key in new_headers:
            if key not in self.headers:
                self.headers.append(key)
                self.ws.cell(row=1, column=self.headers.index(key) + 1, value=key)
                self.get_column(key).auto_size = True
                if pathdict:
                    self.headerpaths.update(pathdict)

    def _write_line(self, line_dict: dict[str, str]) -> None:
        for col_name, raw_value in line_dict.items():
            col_index = self.headers.index(col_name) + 1
            value = self.safe_write_value(raw_value)
            try:
                cell = self.ws.cell(row=self.current_row, column=col_index)
                if isinstance(raw_value, DataItem):
                    cell.value = raw_value.value
                    if isinstance(raw_value.value, int):
                        cell.number_format = f'{self.int_format} "{raw_value.unit}"' if raw_value.unit else self.int_format
                    elif isinstance(raw_value.value, float):
                        cell.number_format = f'{self.float_format} "{raw_value.unit}"' if raw_value.unit else self.float_format
                    else:
                        cell.number_format = '@'  # Text format
                else:
                    self.ws.cell(row=self.current_row, column=col_index, value=self.safe_write_value(raw_value))
                if col_name in self.preset_formats:
                    cell.number_format = self.preset_formats[col_name]
            except Exception as e:
                print(f"Something wrong during writing cell: {e}, raw obj: {raw_value}")
        self.current_row += 1

    def _presave_formatting(self):
        # Raw message is too long, set fixed width
        self.get_column(self.msg_key).width = len(self.msg_key) + 3
        # Frozen the first row and first two columns
        self.ws.freeze_panes = self.frozen_position
        # Set outline according to header paths
        self.ws.sheet_properties.outlinePr.summaryRight = False
        _header: str = self.headers[0]
        for header in self.headers[1:]:
            if _header in self.headerpaths and header in self.headerpaths:
                _path = self.headerpaths[_header]
                path = self.headerpaths[header]
                if len(_path)>4 and len(path)>4:
                    if _path[3]==path[3]:
                        self.get_column(header).outline_level=1
            _header = header

    def safe_write_value(self, value):
        if isinstance(value, (str, int, float, bool, datetime, type(None))):
            return value
        return str(value)
